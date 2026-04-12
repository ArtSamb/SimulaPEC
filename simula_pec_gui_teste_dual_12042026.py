import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import colorsys
from scipy.stats import chi2
import pandas as pd
import time
import threading
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers
from typing import List, Dict, Any, Optional
from numba import jit, prange
from multiprocessing import Pool, cpu_count
import os

os.environ['NUMBA_NUM_THREADS'] = str(cpu_count())

TRANSLATIONS = {
    'pt': {
        'num_pcs_label': 'Número de PCs',
        'erro_pec_label': 'Erro admissível (PEC)',
        'perc_pcs_pec_label': '% PCs acima do PEC',
        'intervalo_acima_label': 'Intervalo acima (%)',
        'intervalo_abaixo_label': 'Intervalo abaixo (%)',
        'num_iteracoes_label': 'Nº de iterações',
        
        'progresso_label': 'Progresso:',
        'progresso_total_label': 'Progresso total:',
        
        'btn_carregar_dados': 'Carregar Dados Reais',
        'btn_confirmar': 'Confirmar',
        'btn_salvar_planilha': 'Salvar Planilha',
        'btn_salvar_grafico': 'Salvar Gráfico',
        'btn_cancelar': 'Cancelar',
        'btn_original': 'Original',
        'btn_english': 'English',
        
        'msg_sucesso': 'Sucesso',
        'msg_erros_carregados': 'erros carregados.',
        'msg_erro': 'Erro',
        'msg_falha_leitura': 'Falha ao ler o arquivo: ',
        'msg_aviso': 'Aviso',
        'msg_executar_simulacao': 'Execute a simulação antes.',
        'msg_exportacao': 'Exportação',
        'msg_planilha_salva': 'Planilha salva com sucesso!',
        'msg_salvo': 'Salvo',
        'msg_grafico_salvo': 'Gráfico salvo com sucesso!',
        'msg_falha_grafico': 'Falha ao salvar o gráfico: ',
        'msg_cancelado': 'Cancelado',
        'msg_simulacao_interrompida': 'A simulação foi interrompida.',
        'msg_erro_inesperado': 'Ocorreu um erro inesperado: ',
        'msg_tempo_estimado': 'Tempo estimado',
        'msg_deseja_continuar': 'Deseja continuar?',
        
        'titulo_precisao': 'Curvas de Rejeição – Precisão',
        'titulo_norma': 'Curvas de Rejeição – Norma do país',
        'xlabel_amostra': 'Tamanho da Amostra',
        'ylabel_prm': 'PRM (%)',
        'legenda_dados_reais': 'R (Dados Reais)',
        'tempo_total': 'Tempo total: ',
        
        'sheet_title': 'Resultado SimulaPEC',
        'tamanho_amostra': 'Tamanho da amostra',
        'prm_precisao': 'PRM(%) - Precisão',
        'prm_norma': 'PRM(%) - Norma',
        'r10': 'R-10',
        
        'lang_title': 'Selecione o idioma / Select Language',
        'lang_instruction': 'Escolha o idioma desejado:',
    },
    'en': {
        'num_pcs_label': 'Number of PCs',
        'erro_pec_label': 'Acceptable error (PEC)',
        'perc_pcs_pec_label': '% PCs above PEC',
        'intervalo_acima_label': 'Interval above (%)',
        'intervalo_abaixo_label': 'Interval below (%)',
        'num_iteracoes_label': 'Nº of iterations',
        
        'progresso_label': 'Progress:',
        'progresso_total_label': 'Total progress:',
        
        'btn_carregar_dados': 'Load Real Data',
        'btn_confirmar': 'Confirm',
        'btn_salvar_planilha': 'Save Spreadsheet',
        'btn_salvar_grafico': 'Save Chart',
        'btn_cancelar': 'Cancel',
        'btn_original': 'Original',
        'btn_english': 'English',
        
        'msg_sucesso': 'Success',
        'msg_erros_carregados': 'errors loaded.',
        'msg_erro': 'Error',
        'msg_falha_leitura': 'Failed to read the file: ',
        'msg_aviso': 'Warning',
        'msg_executar_simulacao': 'Run the simulation before.',
        'msg_exportacao': 'Export',
        'msg_planilha_salva': 'Spreadsheet saved successfully!',
        'msg_salvo': 'Saved',
        'msg_grafico_salvo': 'Chart saved successfully!',
        'msg_falha_grafico': 'Failed to save chart: ',
        'msg_cancelado': 'Canceled',
        'msg_simulacao_interrompida': 'The simulation was interrupted.',
        'msg_erro_inesperado': 'An unexpected error occurred: ',
        'msg_tempo_estimado': 'Estimated time',
        'msg_deseja_continuar': 'Wanna continue?',
        
        'titulo_precisao': 'Rejection Curves – Accuracy',
        'titulo_norma': 'Rejection Curves – Country\'s Standard',
        'xlabel_amostra': 'Sample Size',
        'ylabel_prm': 'PRM (%)',
        'legenda_dados_reais': 'R (Real Data)',
        'tempo_total': 'Total time: ',
        
        'sheet_title': 'Results SimulaPEC',
        'tamanho_amostra': 'Sample Size',
        'prm_precisao': 'PRM(%) - Accuracy',
        'prm_norma': 'PRM(%) - Country\'s Standard',
        'r10': 'R-10',
        
        'lang_title': 'Select Language',
        'lang_instruction': 'Choose your preferred language:',
    }
}


@jit(nopython=True, cache=True)
def gerar_tabela_base(n_pontos, erro_maximo, percentual_acima):
    np.random.seed(0)
    erros = np.zeros(n_pontos, dtype=np.float64)
    
    for i in range(n_pontos):
        u1 = np.random.random()
        u2 = np.random.random()
        erros[i] = np.sqrt(-2.0 * np.log(u1)) * np.cos(2.0 * np.pi * u2)
    
    erros_abs = np.abs(erros)
    erros_ordenados = np.sort(erros_abs)[::-1]
    
    index_limite = int(n_pontos * percentual_acima / 100)
    valor_limite = erros_ordenados[index_limite - 1]
    k = erro_maximo / valor_limite
    
    return erros * k

@jit(nopython=True, parallel=True, cache=True)
def _simular_batch_numba(base, n, n_iter, erro_padrao_precisao, 
                         erro_admissivel, percentual_limite, qui_tabela):
    rejeicoes_p = 0
    rejeicoes_n = 0
    n_base = len(base)
    
    for i in prange(n_iter):
        amostra = np.empty(n, dtype=np.float64)
        for j in range(n):
            idx = np.random.randint(0, n_base)
            amostra[j] = base[idx]
        
        media = np.sum(amostra) / n
        desvio_sq = np.sum((amostra - media) ** 2) / (n - 1)
        
        qui_calc = ((n - 1) * desvio_sq) / (erro_padrao_precisao ** 2)
        if qui_calc > qui_tabela:
            rejeicoes_p += 1
        
        acima = np.sum(np.abs(amostra) > erro_admissivel)
        porcentagem_acima = (acima / n) * 100
        if porcentagem_acima > percentual_limite:
            rejeicoes_n += 1
    
    return rejeicoes_p, rejeicoes_n


def simular_percentual_rejeicao_escalar(
    base, tamanhos, n_iter, erro_admissivel, percentual_limite,
    progress_callback=None, tempo_estimado=False, batch_size=100):
    
    resultado_precisao = []
    resultado_norma = []
    
    start = time.time()
    erro_padrao_precisao = erro_admissivel / 1.6449
    
    qui_tabelas = {n: chi2.ppf(1 - (percentual_limite / 100), df=n - 1) 
                   for n in tamanhos}
    
    if tempo_estimado:
        n = tamanhos[0]
        qui_tabela = qui_tabelas[n]
        tempo_estimado_val = _simular_batch_numba(
            base, n, batch_size, erro_padrao_precisao,
            erro_admissivel, percentual_limite, qui_tabela
        )
        tempo_batch = time.time() - start
        tempo_total_estimado = tempo_batch * (n_iter / batch_size) * len(tamanhos)
        return tempo_total_estimado
    
    for idx, n in enumerate(tamanhos):
        if progress_callback:
            progress_callback(idx + 1, len(tamanhos))
        
        qui_tabela = qui_tabelas[n]
        
        num_batches = (n_iter + batch_size - 1) // batch_size
        rejeicoes_p_total = 0
        rejeicoes_n_total = 0
        
        for batch_idx in range(num_batches):
            current_batch = min(batch_size, n_iter - (batch_idx * batch_size))
            
            rejeicoes_p, rejeicoes_n = _simular_batch_numba(
                base, n, current_batch, erro_padrao_precisao,
                erro_admissivel, percentual_limite, qui_tabela
            )
            
            rejeicoes_p_total += rejeicoes_p
            rejeicoes_n_total += rejeicoes_n
        
        resultado_precisao.append((rejeicoes_p_total / n_iter) * 100)
        resultado_norma.append((rejeicoes_n_total / n_iter) * 100)
    
    return resultado_precisao, resultado_norma

def _salvar_planilha_excel(
    caminho_arquivo: str,
    percentuais: list[float],
    amostras: list[int],
    prm_precisao_list: List[Dict[int, float]],
    prm_norma_list: List[Dict[int, float]],
    dados_reais: bool = False,
    prm_precisao_real: dict[int, float] | None = None,
    prm_norma_real: dict[int, float] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado SimulaPEC"

    col_inicio = 1
    for p in percentuais:
        ws.merge_cells(start_row=1, start_column=col_inicio,
                       end_row=1, end_column=col_inicio + 2)
        cel = ws.cell(row=1, column=col_inicio, value=f"{p}%")
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.font = Font(bold=True)

        ws.cell(row=2, column=col_inicio, value="Tamanho da amostra")
        ws.cell(row=2, column=col_inicio + 1, value="PRM(%) - Precisão")
        ws.cell(row=2, column=col_inicio + 2, value="PRM(%) - Norma")
        col_inicio += 4

    linha_atual = 3
    for n in amostras:
        col_atual = 1
        for i, _perc in enumerate(percentuais):
            dic_prec = prm_precisao_list[i]
            dic_norm = prm_norma_list[i]
            ws.cell(row=linha_atual, column=col_atual, value=n)
            ws.cell(row=linha_atual, column=col_atual + 1,
                    value=round(dic_prec.get(n, 0.0), 2))
            ws.cell(row=linha_atual, column=col_atual + 2,
                    value=round(dic_norm.get(n, 0.0), 2))
            for c in (col_atual + 1, col_atual + 2):
                ws.cell(row=linha_atual, column=c).number_format = \
                    numbers.FORMAT_NUMBER_00
            col_atual += 4
        linha_atual += 1

    if dados_reais:
        linha_atual += 1
        ws.merge_cells(start_row=linha_atual, start_column=1,
                       end_row=linha_atual, end_column=3)
        cel = ws.cell(row=linha_atual, column=1, value="R-10")
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.font = Font(bold=True)
        ws.cell(row=linha_atual + 1, column=1, value="Tamanho da amostra")
        ws.cell(row=linha_atual + 1, column=2, value="PRM(%) - Precisão")
        ws.cell(row=linha_atual + 1, column=3, value="PRM(%) - Norma")
        linha_dados = linha_atual + 2
        for n in amostras:
            ws.cell(row=linha_dados, column=1, value=n)
            ws.cell(row=linha_dados, column=2,
                    value=round(prm_precisao_real.get(n, 0.0), 2))
            ws.cell(row=linha_dados, column=3,
                    value=round(prm_norma_real.get(n, 0.0), 2))
            for c in (2, 3):
                ws.cell(row=linha_dados, column=c).number_format = \
                    numbers.FORMAT_NUMBER_00
            linha_dados += 1

    for i, coluna in enumerate(ws.columns, start=1):
        max_larg = max((len(str(cel.value)) for cel in coluna if cel.value),
                       default=0)

        ws.column_dimensions[get_column_letter(i)].width = max_larg + 2
    wb.save(caminho_arquivo)

DEFAULT_LINEWIDTH = 2.5
def _line_width():
    return DEFAULT_LINEWIDTH


class SimulaPECApp:
    def __init__(self, master, lang='pt'):
        self.master = master
        self.lang = lang
        self.t = TRANSLATIONS[lang]
        
        master.title(f"SimulaPEC {'Teste' if lang == 'pt' else 'Test'} 12/04/2026")

        self.dados_reais = None
        self.curvas_precisao = []
        self.curvas_norma = []
        self.curvas_precisao: List[List[float]] = []
        self.curvas_norma: List[List[float]] = []
        self.percentuais = []
        self.tamanhos_amostra = []
        self.curva_precisao_R = None
        self.curva_norma_R = None
        self.figura = None
        self.entries = {}
        self.perc_max = None
        self.intervalo = None
        self.amostras = []
        self.prm_precisao = {}
        self.prm_norma = {}
        self.dados_reais = False
        self.prm_precisao_real = {}
        self.prm_norma_real = {}
        self.prm_precisao_real: Dict[int, float] = {}
        self.prm_norma_real: Dict[int, float] = {}
        self.prm_precisao_list: List[Dict[int, float]] = []
        self.prm_norma_list: List[Dict[int, float]] = []

        self._cancel_requested = False
        self._thread_worker = None

        campos = [
            (self.t['num_pcs_label'], "150"),
            (self.t['erro_pec_label'], "5"),
            (self.t['perc_pcs_pec_label'], "10"),
            (self.t['intervalo_acima_label'], "5"),
            (self.t['intervalo_abaixo_label'], "2"),
            (self.t['num_iteracoes_label'], "3000")]

        for i, (label, default) in enumerate(campos):
            tk.Label(master, text=label).grid(row=i+2, column=0, sticky="e")
            entry = tk.Entry(master)
            entry.insert(0, default)
            entry.grid(row=i+2, column=1)
            self.entries[label] = entry

        self.label_progresso = tk.Label(master, text=f"{self.t['progresso_label']} 0%")
        self.label_progresso.grid(row=9, column=0, columnspan=2)

        self.progress = ttk.Progressbar(master, length=200, mode='determinate')
        self.progress.grid(row=10, column=0, columnspan=2, pady=5)

        self.label_total = tk.Label(master, text=f"{self.t['progresso_total_label']} 0%")
        self.label_total.grid(row=11, column=0, columnspan=2)

        self.progress_total = ttk.Progressbar(master, length=200,
                                              mode='determinate')
        self.progress_total.grid(row=12, column=0, columnspan=2, pady=5)

        tk.Button(master, text=self.t['btn_carregar_dados'],
                  command=self.carregar_dados_reais).grid(row=13, column=0)
        tk.Button(master, text=self.t['btn_confirmar'],
                  command=self.executar).grid(row=13, column=1)
        tk.Button(master, text=self.t['btn_salvar_planilha'],
                  command=self.exportar_planilha).grid(row=14, column=0)
        tk.Button(master, text=self.t['btn_salvar_grafico'],
                  command=self.exportar_grafico).grid(row=14, column=1)

        self.btn_cancelar = tk.Button(master, text=self.t['btn_cancelar'],
                                      command=self.cancelar, state="disabled")
        self.btn_cancelar.grid(row=15, column=0, columnspan=2, pady=5)

    def disable_ui_during_run(self):
        self.progress_total['value'] = 0
        self.label_total.config(text=f"{self.t['progresso_total_label']} 0%")
        for child in self.master.winfo_children():
            if isinstance(child, tk.Button) and child["text"] != self.t['btn_cancelar']:
                child.configure(state="disabled")
        self.btn_cancelar.configure(state="normal")
        self.progress['value'] = 0
        self.label_progresso.config(text=f"{self.t['progresso_label']} 0%")

    def enable_ui_after_run(self):
        for child in self.master.winfo_children():
            if isinstance(child, tk.Button):
                child.configure(state="normal")

        self.btn_cancelar.configure(state="disabled")
        self.progress_total['value'] = 100
        self.label_total.config(text=f"{self.t['progresso_total_label']} 100%")

    def cancelar(self):
        self._cancel_requested = True
        self.btn_cancelar.configure(state="disabled")

    def atualizar_progresso(self, valor, total):
        pct = int((valor / total) * 100)
        self.progress['value'] = pct
        self.label_progresso.config(text=f"{self.t['progresso_label']} {pct}%")
        self.master.update_idletasks()

    def atualizar_progresso_total(self, concluido, total):
        pct = int((concluido / total) * 100)
        self.progress_total['value'] = pct
        self.label_total.config(text=f"{self.t['progresso_total_label']} {pct}%")
        self.master.update_idletasks()

    def _gerar_cores_personalizadas(self, percentuais, perc_base):
        cores = []

        dist_max = max(
            abs(p - perc_base) for p in percentuais if p != perc_base
        ) if len([p for p in percentuais if p != perc_base]) > 0 else 1

        for perc in percentuais:
            if perc == perc_base:
                cores.append('#1f77b4')
            elif perc < perc_base:
                distancia = abs(perc - perc_base) / dist_max
                r, g, b = colorsys.hls_to_rgb(0.33, 0.9 - (0.6 * distancia), 0.5)
                cores.append((r, g, b))
            else:
                distancia = abs(perc - perc_base) / dist_max
                r, g, b = colorsys.hls_to_rgb(0.0, 0.9 - (0.6 * distancia), 0.5)
                cores.append((r, g, b))
        return cores

    def carregar_dados_reais(self):
        caminho = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if caminho:
            try:
                with open(caminho, "r") as f:
                    linhas = f.readlines()
                    self.dados_reais = [
                        float(l.strip().replace(",", "."))
                        for l in linhas if l.strip()
                    ]
                messagebox.showinfo(self.t['msg_sucesso'],
                                    f"{len(self.dados_reais)} {self.t['msg_erros_carregados']}")
            except Exception as e:
                messagebox.showerror(self.t['msg_erro'], 
                                     f"{self.t['msg_falha_leitura']}{e}")

    def executar(self):
        self._cancel_requested = False
        self.disable_ui_during_run()
        self._thread_worker = threading.Thread(target=self._processar_simulacao)
        self._thread_worker.start()

    def _ask_continue(self, tempo_total):
        self._continue_resp = None

        def _callback():
            resp = messagebox.askyesno(
                self.t['msg_tempo_estimado'],
                f"{int(tempo_total)} s.\n{self.t['msg_deseja_continuar']}"
            )
            self._continue_resp = resp

        self.master.after(0, _callback)

        while self._continue_resp is None:
            self.master.update()
            time.sleep(0.05)
        return self._continue_resp

    def _finalizar_thread(self, cancelado: bool):
        def _restore():
            self.enable_ui_after_run()
            if cancelado:
                messagebox.showinfo(self.t['msg_cancelado'], 
                                    self.t['msg_simulacao_interrompida'])
        self.master.after(0, _restore)

    def _processar_simulacao(self):
        try:
            N = int(self.entries[self.t['num_pcs_label']].get())
            erro_admissivel = float(self.entries[self.t['erro_pec_label']].get())
            perc_base = float(self.entries[self.t['perc_pcs_pec_label']].get())
            self.perc_base = perc_base
            intervalo_acima = float(self.entries[self.t['intervalo_acima_label']].get())
            intervalo_abaixo = float(self.entries[self.t['intervalo_abaixo_label']].get())
            n_iter = int(self.entries[self.t['num_iteracoes_label']].get())

            if N <= 0:
                raise ValueError()
            if erro_admissivel <= 0:
                raise ValueError()
            if perc_base <= 0:
                raise ValueError()
            if intervalo_acima <= 0:
                raise ValueError()
            if intervalo_abaixo <= 0:
                raise ValueError()
            if n_iter <= 0:
                raise ValueError()

            self.tamanhos_amostra = list(range(5, int(N * 0.6) + 1, 5))
            self.percentuais = []

            perc = perc_base - intervalo_abaixo
            while perc > 0:
                self.percentuais.append(perc)
                perc -= intervalo_abaixo
            self.percentuais.append(perc_base)

            perc = perc_base + intervalo_acima
            while perc <= 40:
                self.percentuais.append(perc)
                perc += intervalo_acima

            self.percentuais.sort()
            base_teste = gerar_tabela_base(N, erro_admissivel, perc_base)
            tempo_unit = simular_percentual_rejeicao_escalar(
                base_teste, self.tamanhos_amostra, n_iter,
                erro_admissivel, perc_base, tempo_estimado=True)
            tempo_total = tempo_unit * len(self.percentuais)

            if not self._ask_continue(tempo_total):
                self._finalizar_thread(cancelado=True)
                return

            self.inicio_processamento = time.time()
            self.curvas_precisao.clear()
            self.curvas_norma.clear()
            self.prm_precisao_list.clear()
            self.prm_norma_list.clear()
            self.prm_precisao_real.clear()
            self.prm_norma_real.clear()

            if N > 200000:
                batch_size = 50
            elif N > 100000:
                batch_size = 100
            else:
                batch_size = 200

            for idx, perc in enumerate(self.percentuais):
                if self._cancel_requested:
                    pool.terminate()
                    self._finalizar_thread(cancelado=True)
                    return

                base = gerar_tabela_base(N, erro_admissivel, perc)
                if N > 100000:
                    base = base.astype(np.float32)
                pr_p, pr_n = simular_percentual_rejeicao_escalar(
                    base, self.tamanhos_amostra, n_iter,
                    erro_admissivel, perc,
                    progress_callback=self.atualizar_progresso,
                    batch_size=batch_size)
                self.curvas_precisao.append(pr_p)
                self.curvas_norma.append(pr_n)
                dic_prec = {n: round(v, 2) for n, v in zip(self.tamanhos_amostra,
                                                           pr_p)}
                dic_norm = {n: round(v, 2) for n, v in zip(self.tamanhos_amostra,
                                                           pr_n)}
                self.prm_precisao_list.append(dic_prec)
                self.prm_norma_list.append(dic_norm)

                self.atualizar_progresso_total(
                    concluido=idx + 1,
                    total=len(self.percentuais))

            if self.dados_reais:
                base_real = np.array(self.dados_reais)
                if len(base_real) > 100000:
                    base_real = base_real.astype(np.float32)
                pr_p_r, pr_n_r = simular_percentual_rejeicao_escalar(
                    base_real, self.tamanhos_amostra, n_iter,
                    erro_admissivel, perc_base,
                    batch_size=batch_size)

                self.curva_precisao_R = pr_p_r
                self.curva_norma_R = pr_n_r
                self.prm_precisao_real = {
                    n: round(v, 2) for n, v in zip(self.tamanhos_amostra, pr_p_r)}
                self.prm_norma_real = {
                    n: round(v, 2) for n, v in zip(self.tamanhos_amostra, pr_n_r)}
                self.atualizar_progresso_total(
                    concluido=len(self.percentuais) + 1,
                    total=len(self.percentuais) + 1)

            self.tempo_processamento = time.time() - self.inicio_processamento
            self.master.after(0, self.plotar)
            self._finalizar_thread(cancelado=False)

        except Exception as exc:
            self.master.after(0, lambda: messagebox.showerror(
                self.t['msg_erro'], f"{self.t['msg_erro_inesperado']}{exc}"))
            self._finalizar_thread(cancelado=True)

    def plotar(self):
        fig, ax = plt.subplots(1, 2, figsize=(14, 6))
        perc_base = self.perc_base
        n_curvas = len(self.percentuais)
        cores = self._gerar_cores_personalizadas(self.percentuais, perc_base)

        for i, (perc, curva) in enumerate(zip(self.percentuais,
                                              self.curvas_precisao)):
            ax[0].plot(self.tamanhos_amostra, curva,
                       label=f"{perc}%", color=cores[i], linewidth=_line_width())

        if self.curva_precisao_R is not None:
            ax[0].plot(self.tamanhos_amostra, self.curva_precisao_R,
                       label=self.t['legenda_dados_reais'], 
                       linewidth=_line_width() * 1.5,
                       color='black', linestyle='--')

        ax[0].set_title(self.t['titulo_precisao'])
        ax[0].set_xlabel(self.t['xlabel_amostra'])
        ax[0].set_ylabel(self.t['ylabel_prm'])
        ax[0].legend()
        ax[0].grid(True)

        for i, (perc, curva) in enumerate(zip(self.percentuais,
                                              self.curvas_norma)):
            ax[1].plot(self.tamanhos_amostra, curva,
                       label=f"{perc}%", color=cores[i], linewidth=_line_width())

        if self.curva_norma_R is not None:
            ax[1].plot(self.tamanhos_amostra, self.curva_norma_R,
                       label=self.t['legenda_dados_reais'], 
                       linewidth=_line_width() * 1.5,
                       color='black', linestyle='--')

        ax[1].set_title(self.t['titulo_norma'])
        ax[1].set_xlabel(self.t['xlabel_amostra'])
        ax[1].set_ylabel(self.t['ylabel_prm'])
        ax[1].legend()
        ax[1].grid(True)

        plt.tight_layout()
        ax[0].xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        ax[1].xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        secs = self.tempo_processamento
        mins, secs = divmod(secs, 60)

        tempo_str = f"{int(mins)} min {secs:.2f} s" if mins else f"{secs:.2f} s"

        fig.text(0.58, 0.98,
                 f"{self.t['tempo_total']}{tempo_str}",
                 ha='right', va='top',
                 fontsize=10,
                 bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))

        plt.show(block=False)
        self.figura = fig

    def exportar_planilha(self):
        if not self.curvas_precisao:
            messagebox.showwarning(self.t['msg_aviso'], self.t['msg_executar_simulacao'])
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not caminho:
            return

        amostras = self.tamanhos_amostra

        _salvar_planilha_excel(
            caminho_arquivo=caminho,
            percentuais=self.percentuais,
            amostras=amostras,
            prm_precisao_list=self.prm_precisao_list,
            prm_norma_list=self.prm_norma_list,
            dados_reais=self.dados_reais,
            prm_precisao_real=self.prm_precisao_real,
            prm_norma_real=self.prm_norma_real,)

        messagebox.showinfo(self.t['msg_exportacao'], self.t['msg_planilha_salva'])

    def exportar_grafico(self):
        if not self.figura:
            messagebox.showwarning(self.t['msg_aviso'], self.t['msg_executar_simulacao'])
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("Imagem PNG", "*.png")]
        )
        if not caminho:
            return

        try:
            self.figura.savefig(caminho, dpi=300, bbox_inches="tight")
            messagebox.showinfo(self.t['msg_salvo'], self.t['msg_grafico_salvo'])
        except Exception as exc:
            messagebox.showerror(self.t['msg_erro'], f"{self.t['msg_falha_grafico']}{exc}")


class LanguageSelector:
    def __init__(self, root):
        self.root = root
        self.root.title("SimulaPEC - Selecione o idioma / Select Language")
        self.root.geometry("400x200")
        self.root.resizable(False, False)
        self.selected_lang = None
        
        self.root.update_idletasks()
        width = 400
        height = 200
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        frame = tk.Frame(root, padx=20, pady=20)
        frame.pack(expand=True)
        
        tk.Label(frame, text="Escolha o idioma / Choose language:", 
                 font=('Arial', 12, 'bold'), pady=10).pack()
        
        btn_frame = tk.Frame(frame)
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Original", width=20, height=2,
                  bg="#e1f5fe", command=lambda: self.select_lang('pt')).pack(side=tk.LEFT, padx=10)
        
        tk.Button(btn_frame, text="English", width=20, height=2,
                  bg="#e1f5fe", command=lambda: self.select_lang('en')).pack(side=tk.LEFT, padx=10)
    
    def select_lang(self, lang):
        self.selected_lang = lang
        self.root.withdraw()
        self.root.event_generate("<<LangSelected>>", when="tail")


def main():
    root = tk.Tk()
    selector = LanguageSelector(root)
    app_instance = [None] 
    
    def on_lang_selected(event):
        if selector.selected_lang:
            selector.root.destroy()
            
            new_root = tk.Tk()
            new_root.title(f"SimulaPEC {'Teste' if selector.selected_lang == 'pt' else 'Test'} 12/04/2026")
            
            app_instance[0] = SimulaPECApp(new_root, lang=selector.selected_lang)
            
            new_root.mainloop()
    
    root.bind("<<LangSelected>>", on_lang_selected)
    
    root.mainloop()

if __name__ == "__main__":
    main()