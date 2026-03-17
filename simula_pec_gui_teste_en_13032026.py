import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import colorsys                     
from scipy.stats import chi2
import pandas as pd
import time
import threading
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers
from typing import List, Dict, Any

def gerar_tabela_base(n_pontos, erro_maximo, percentual_acima):
    np.random.seed(0)
    erros = np.random.normal(loc=0, scale=1, size=n_pontos)
    erros_abs = np.abs(erros)
    erros_ordenados = np.sort(erros_abs)[::-1]

    index_limite = int(n_pontos * percentual_acima / 100)
    valor_limite = erros_ordenados[index_limite - 1]
    k = erro_maximo / valor_limite
    return erros * k

def teste_precisao(amostra, erro_padrao_aceitavel, percentual_limite):
    n = len(amostra)
    desvio_padrao = np.std(amostra, ddof=1)
    qui_calc = ((n - 1) * (desvio_padrao ** 2)) / (erro_padrao_aceitavel ** 2)
    qui_tabela = chi2.ppf(1 - (percentual_limite / 100), df=n - 1)
    return qui_calc <= qui_tabela

def teste_norma_pais(amostra, erro_admissivel, percentual_limite): 
    acima = np.sum(np.abs(amostra) > erro_admissivel)
    porcentagem_acima = (acima / len(amostra)) * 100
    return porcentagem_acima <= percentual_limite

def simular_percentual_rejeicao(
    base, tamanhos, n_iter, erro_admissivel, percentual_limite,
    progress_callback=None, tempo_estimado=False):
    resultado_precisao = []
    resultado_norma = []
    start = time.time()

    for idx, n in enumerate(tamanhos):
        rejeicoes_p = rejeicoes_n = 0

        for _ in range(n_iter):
            amostra = np.random.choice(base, size=n, replace=True)

            if not teste_precisao(amostra, erro_admissivel / 1.6449, percentual_limite):
                rejeicoes_p += 1

            if not teste_norma_pais(amostra, erro_admissivel, percentual_limite):
                rejeicoes_n += 1

        resultado_precisao.append((rejeicoes_p / n_iter) * 100)
        resultado_norma.append((rejeicoes_n / n_iter) * 100)

        if progress_callback:
            progress_callback(idx + 1, len(tamanhos))

        if tempo_estimado:
            break

    if tempo_estimado:
        return time.time() - start

    return resultado_precisao, resultado_norma

def _salvar_planilha_excel(
    caminho_arquivo: str,
    percentuais: list[float],
    amostras: list[int],
    prm_precisao_list: List[Dict[int, float]],
    prm_norma_list: List[Dict[int, float]],
    dados_reais: bool = False,
    prm_precisao_real: dict[int, float] | None = None,
    prm_norma_real: dict[int, float] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado SimulaPEC"

    col_inicio = 1
    for p in percentuais:
        ws.merge_cells(start_row=1, start_column=col_inicio,
                       end_row=1, end_column=col_inicio + 2)
        cel = ws.cell(row=1, column=col_inicio, value=f"{p}%")
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.font = Font(bold=True)

        ws.cell(row=2, column=col_inicio, value="Tamanho da amostra")
        ws.cell(row=2, column=col_inicio + 1, value="PRM(%) - Precisão")
        ws.cell(row=2, column=col_inicio + 2, value="PRM(%) - Norma")
        col_inicio += 4

    linha_atual = 3
    for n in amostras:
        col_atual = 1
        for i, _perc in enumerate(percentuais):
            dic_prec = prm_precisao_list[i]
            dic_norm = prm_norma_list[i]
            ws.cell(row=linha_atual, column=col_atual, value=n)
            ws.cell(row=linha_atual, column=col_atual + 1,
                    value=round(dic_prec.get(n, 0.0), 2))
            ws.cell(row=linha_atual, column=col_atual + 2,
                    value=round(dic_norm.get(n, 0.0), 2))
            for c in (col_atual + 1, col_atual + 2):
                ws.cell(row=linha_atual, column=c).number_format = numbers.FORMAT_NUMBER_00
            col_atual += 4
        linha_atual += 1

    if dados_reais:
        linha_atual += 1
        ws.merge_cells(start_row=linha_atual, start_column=1,
                       end_row=linha_atual, end_column=3)
        cel = ws.cell(row=linha_atual, column=1, value="R-10")
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.font = Font(bold=True)
        ws.cell(row=linha_atual + 1, column=1, value="Tamanho da amostra")
        ws.cell(row=linha_atual + 1, column=2, value="PRM(%) - Precisão")
        ws.cell(row=linha_atual + 1, column=3, value="PRM(%) - Norma")
        linha_dados = linha_atual + 2
        for n in amostras:
            ws.cell(row=linha_dados, column=1, value=n)
            ws.cell(row=linha_dados, column=2,
                    value=round(prm_precisao_real.get(n, 0.0), 2))
            ws.cell(row=linha_dados, column=3,
                    value=round(prm_norma_real.get(n, 0.0), 2))
            for c in (2, 3):
                ws.cell(row=linha_dados, column=c).number_format = numbers.FORMAT_NUMBER_00
            linha_dados += 1

    for i, coluna in enumerate(ws.columns, start=1):
        max_larg = max((len(str(cel.value)) for cel in coluna if cel.value), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max_larg + 2
    wb.save(caminho_arquivo)

DEFAULT_LINEWIDTH = 2.5
def _line_width():
    return DEFAULT_LINEWIDTH

class SimulaPECApp:
    def __init__(self, master):
        self.master = master
        master.title("SimulaPEC Test 13/03/2026")

        self.dados_reais = None
        self.curvas_precisao = []
        self.curvas_norma = []
        self.curvas_precisao: List[List[float]] = []
        self.curvas_norma: List[List[float]] = []
        self.percentuais = []
        self.tamanhos_amostra = []
        self.curva_precisao_R = None
        self.curva_norma_R = None
        self.figura = None
        self.entries = {}
        self.perc_max = None
        self.intervalo = None
        self.amostras = []
        self.prm_precisao = {}
        self.prm_norma = {}
        self.dados_reais = False
        self.prm_precisao_real = {}
        self.prm_norma_real = {}
        self.prm_precisao_real: Dict[int, float] = {}
        self.prm_norma_real: Dict[int, float] = {}
        self.prm_precisao_list: List[Dict[int, float]] = []
        self.prm_norma_list: List[Dict[int, float]] = []

        self._cancel_requested = False
        self._thread_worker = None

        campos = [
            ("Number of PCs", "150"),
            ("Acceptable error (PEC)", "5"),
            ("% PCs above PEC", "10"),
            ("Interval above (%)", "5"),
            ("Interval below (%)", "2"),
            ("Nº of iterations", "3000")]

        for i, (label, default) in enumerate(campos):
            tk.Label(master, text=label).grid(row=i, column=0, sticky="e")
            entry = tk.Entry(master)
            entry.insert(0, default)
            entry.grid(row=i, column=1)
            self.entries[label] = entry

        self.label_progresso = tk.Label(master, text="Progress: 0%")
        self.label_progresso.grid(row=7, column=0, columnspan=2)

        self.progress = ttk.Progressbar(master, length=200, mode='determinate')
        self.progress.grid(row=8, column=0, columnspan=2, pady=5)

        self.label_total = tk.Label(master, text="Total progress: 0%")
        self.label_total.grid(row=9, column=0, columnspan=2)

        self.progress_total = ttk.Progressbar(master, length=200, mode='determinate')
        self.progress_total.grid(row=10, column=0, columnspan=2, pady=5)

        tk.Button(master, text="Load Real Data",
                  command=self.carregar_dados_reais).grid(row=11, column=0)
        tk.Button(master, text="Confirm",
                  command=self.executar).grid(row=11, column=1)
        tk.Button(master, text="Save Spreadsheet",
                  command=self.exportar_planilha).grid(row=12, column=0)
        tk.Button(master, text="Save Chart",
                  command=self.exportar_grafico).grid(row=12, column=1)

        self.btn_cancelar = tk.Button(master, text="Cancel",
                                      command=self.cancelar, state="disabled")
        self.btn_cancelar.grid(row=13, column=0, columnspan=2, pady=5)

    def disable_ui_during_run(self):
        self.progress_total['value'] = 0
        self.label_total.config(text="Total progress: 0%")
        for child in self.master.winfo_children():
            if isinstance(child, tk.Button) and child["text"] != "Cancel":
                child.configure(state="disabled")
        self.btn_cancelar.configure(state="normal")
        self.progress['value'] = 0
        self.label_progresso.config(text="Progress: 0%")

    def enable_ui_after_run(self):
        for child in self.master.winfo_children():
            if isinstance(child, tk.Button):
                child.configure(state="normal")
        self.btn_cancelar.configure(state="disabled")
        self.progress_total['value'] = 100
        self.label_total.config(text="Total progess: 100%")

    def cancelar(self):
        self._cancel_requested = True
        self.btn_cancelar.configure(state="disabled")

    def atualizar_progresso(self, valor, total):
        pct = int((valor / total) * 100)
        self.progress['value'] = pct
        self.label_progresso.config(text=f"Progress: {pct}%")
        self.master.update_idletasks()

    def atualizar_progresso_total(self, concluido, total):
        pct = int((concluido / total) * 100)
        self.progress_total['value'] = pct
        self.label_total.config(text=f"Total progress: {pct}%")
        self.master.update_idletasks()

    def _gerar_cores(self, n):
        cmap = plt.get_cmap('tab20')

        def escurecer(rgba, fator=0.7):
            r, g, b, a = rgba
            h, l, s = colorsys.rgb_to_hls(r, g, b)
            l = max(0, min(1, l * fator))
            r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
            return (r2, g2, b2, a)

        return [escurecer(cmap(i / max(n, 1))) for i in range(n)]

    def carregar_dados_reais(self):
        caminho = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if caminho:
            try:
                with open(caminho, "r") as f:
                    linhas = f.readlines()
                self.dados_reais = [
                    float(l.strip().replace(",", "."))
                    for l in linhas if l.strip()
                ]
                messagebox.showinfo("Success",
                                    f"{len(self.dados_reais)} errors loaded.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read the file: {e}")

    def executar(self):
        self._cancel_requested = False
        self.disable_ui_during_run()
        self._thread_worker = threading.Thread(target=self._processar_simulacao)
        self._thread_worker.start()

    def _ask_continue(self, tempo_total):
        self._continue_resp = None

        def _callback():
            resp = messagebox.askyesno(
                "Estimated time",
                f"Estimated time: {int(tempo_total)} s.\nWanna continue?"
            )
            self._continue_resp = resp

        self.master.after(0, _callback)

        while self._continue_resp is None:
            time.sleep(0.05)
        return self._continue_resp

    def _finalizar_thread(self, cancelado: bool):
        def _restore():
            self.enable_ui_after_run()
            if cancelado:
                messagebox.showinfo("Canceled", "The simulation was interrupted.")
        self.master.after(0, _restore)
    
    def iniciar_simulacao(self):
        self.perc_max   = int(self.entry_perc_max.get())
        self.intervalo  = int(self.entry_intervalo.get())

    def _processar_simulacao(self):
        try:
            N = int(self.entries["Number of PCs"].get())
            erro_admissivel = float(self.entries["Acceptable error (PEC)"].get())
            perc_base = float(self.entries["% PCs above PEC"].get())
            intervalo_acima = float(self.entries["Interval above (%)"].get())
            intervalo_abaixo = float(self.entries["Interval below (%)"].get())
            n_iter = int(self.entries["Nº of iterations"].get())

            if N <= 0:
                raise ValueError()
            if erro_admissivel <= 0:
                raise ValueError()
            if perc_base <= 0:
                raise ValueError()
            if intervalo_acima <= 0:
                raise ValueError()
            if intervalo_abaixo <= 0:
                raise ValueError()
            if n_iter <= 0:
                raise ValueError()

            self.tamanhos_amostra = list(range(5, int(N * 0.6) + 1, 5))
            self.percentuais = []

            perc = perc_base - intervalo_abaixo
            while perc > 0:
                self.percentuais.append(perc)
                perc -= intervalo_abaixo
            self.percentuais.append(perc_base)

            perc = perc_base + intervalo_acima
            while perc <= 40:
                self.percentuais.append(perc)
                perc += intervalo_acima
            
            self.percentuais.sort()
            base_teste = gerar_tabela_base(N, erro_admissivel, perc_base)
            tempo_unit = simular_percentual_rejeicao(
                base_teste, self.tamanhos_amostra, n_iter,
                erro_admissivel, perc_base, tempo_estimado=True)
            tempo_total = tempo_unit * len(self.percentuais)

            if not self._ask_continue(tempo_total):
                self._finalizar_thread(cancelado=True)
                return
            
            self.inicio_processamento = time.time()
            self.curvas_precisao.clear()
            self.curvas_norma.clear()
            self.prm_precisao_list.clear()
            self.prm_norma_list.clear()
            self.prm_precisao_real.clear()
            self.prm_norma_real.clear()

            for idx, perc in enumerate(self.percentuais):
                if self._cancel_requested:
                    self._finalizar_thread(cancelado=True)
                    return
                
                base = gerar_tabela_base(N, erro_admissivel, perc)
                pr_p, pr_n = simular_percentual_rejeicao(
                    base, self.tamanhos_amostra, n_iter,
                    erro_admissivel, perc,
                    progress_callback=self.atualizar_progresso)
                self.curvas_precisao.append(pr_p)
                self.curvas_norma.append(pr_n)
                dic_prec = {n: round(v, 2) for n, v in zip(self.tamanhos_amostra, pr_p)}
                dic_norm = {n: round(v, 2) for n, v in zip(self.tamanhos_amostra, pr_n)}
                self.prm_precisao_list.append(dic_prec)
                self.prm_norma_list.append(dic_norm)

                self.atualizar_progresso_total(
                    concluido=idx + 1,
                    total=len(self.percentuais))
            
            if self.dados_reais:
                base_real = np.array(self.dados_reais)
                pr_p_r, pr_n_r = simular_percentual_rejeicao(
                    base_real, self.tamanhos_amostra, n_iter,
                    erro_admissivel, perc_base)
                
                self.curva_precisao_R = pr_p_r
                self.curva_norma_R = pr_n_r
                self.prm_precisao_real = {
                    n: round(v, 2) for n, v in zip(self.tamanhos_amostra, pr_p_r)}
                self.prm_norma_real = {
                    n: round(v, 2) for n, v in zip(self.tamanhos_amostra, pr_n_r)}
                self.atualizar_progresso_total(
                    concluido=len(self.percentuais) + 1,
                    total=len(self.percentuais) + 1)
            
            self.tempo_processamento = time.time() - self.inicio_processamento
            self.master.after(0, self.plotar)
            self._finalizar_thread(cancelado=False)
        
        except Exception as exc:
            self.master.after(0, lambda: messagebox.showerror(
                "Error", f"An unexpected error occurred: {exc}"))
            self._finalizar_thread(cancelado=True)

    def plotar(self):
        fig, ax = plt.subplots(1, 2, figsize=(14, 6))
        n_curvas = len(self.percentuais)
        cores = self._gerar_cores(n_curvas)

        for i, (perc, curva) in enumerate(zip(self.percentuais, self.curvas_precisao)):
            ax[0].plot(self.tamanhos_amostra, curva,
                       label=f"{perc}%", color=cores[i], linewidth=_line_width())
        
        if self.curva_precisao_R is not None:
            ax[0].plot(self.tamanhos_amostra, self.curva_precisao_R,
                       label="R (Dados Reais)", linewidth=_line_width() * 1.5, color='red', linestyle='--')
        
        ax[0].set_title("Rejection Curves – Accuracy")
        ax[0].set_xlabel("Sample Size")
        ax[0].set_ylabel("PRM (%)")
        ax[0].legend()
        ax[0].grid(True)

        for i, (perc, curva) in enumerate(zip(self.percentuais, self.curvas_norma)):
            ax[1].plot(self.tamanhos_amostra, curva,
                       label=f"{perc}%", color=cores[i], linewidth=_line_width())
        
        if self.curva_norma_R is not None:
            ax[1].plot(self.tamanhos_amostra, self.curva_norma_R,
                       label="R (Dados Reais)", linewidth=_line_width() * 1.5, color='red', linestyle='--')

        ax[1].set_title("Rejection Curves – Country's Standard")
        ax[1].set_xlabel("Sample Size")
        ax[1].set_ylabel("PRM (%)")
        ax[1].legend()
        ax[1].grid(True)

        plt.tight_layout()
        ax[0].xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        ax[1].xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        secs = self.tempo_processamento
        mins, secs = divmod(secs, 60)
        tempo_str = f"{int(mins)} min {secs:.2f} s" if mins else f"{secs:.2f} s"

        fig.text(0.58, 0.98,
                 f"Total time: {tempo_str}",
                 ha='right', va='top',
                 fontsize=10,
                 bbox=dict(facecolor='white', alpha=0.7, edgecolor='gray'))

        plt.show(block=False)
        self.figura = fig

    def exportar_planilha(self):
        if not self.curvas_precisao:
            messagebox.showwarning("Warning", "Run the simulation before.")
            return
        
        caminho = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
        if not caminho:
            return
        
        amostras = self.tamanhos_amostra

        _salvar_planilha_excel(
            caminho_arquivo=caminho,
            percentuais=self.percentuais,
            amostras=amostras,
            prm_precisao_list=self.prm_precisao_list,
            prm_norma_list=self.prm_norma_list,
            dados_reais=self.dados_reais,
            prm_precisao_real=self.prm_precisao_real,
            prm_norma_real=self.prm_norma_real,)
        
        messagebox.showinfo("Exportação", "Planilha salva com sucesso!")

    def exportar_grafico(self):
        if not self.figura:
            messagebox.showwarning("Warning", "Run the simulation before.")
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("Imagem PNG", "*.png")]
        )
        if not caminho:
            return

        try:
            self.figura.savefig(caminho, dpi=300, bbox_inches="tight")
            messagebox.showinfo("Saved", "Chart saved successfully!")
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to save chart: {exc}")


if __name__ == "__main__":
    root = tk.Tk()
    app = SimulaPECApp(root)
    root.mainloop()